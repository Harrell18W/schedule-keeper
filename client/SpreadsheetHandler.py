from calendar import monthrange
import datetime as dt
import openpyxl as pyxl

class SpreadsheetHandler:

    def __init__(self, filename: str, engineer_columns: list, active_sheet: str="-1", title_row: int=1, date_column:int=1):
        self.filename = filename
        self.workbook = pyxl.load_workbook(filename)
        if active_sheet == "-1":
            self.active_sheet = self.workbook[dt.datetime(2000, 1, 1).now().strftime("%Y")]
        else:
            self.active_sheet = self.workbook[active_sheet]
        self.engineer_columns = engineer_columns
        self.title_row = title_row
        self.date_column = date_column

    #save the workbook
    def save(self):
        self.workbook.save(self.filename)

    #list all engineers
    def engineer_names(self):
        names = []
        for engineer in engineer_columns:
            names.append(self.active_sheet[engineer + str(self.title_row)].value)
        names.sort()
        return names

    #list all customers
    def customer_names(self):
        names = []
        for column in range(1, self.active_sheet.max_column + 1):
            value = self.active_sheet.cell(row=self.title_row, column=column).value
            if column == self.title_row or \
                pyxl.utils.get_column_letter(column) in self.engineer_columns or \
                value in names:
                continue
            names.append(value)
        names.sort()
        return names

    #get row with corresponding date
    def row_from_date(self, date: dt.datetime):
        for row in range(1, self.active_sheet.max_row + 1):
            value = self.active_sheet.cell(row=row, column=self.date_column).value
            if not isinstance(value, dt.datetime):
                continue
            if value == date:
                return row
        return -1

    #create row with corresponding date
    def new_row(self, date: dt.date):
        last_date_row: int = -1
        for row in range(1, self.active_sheet.max_row + 1):
            if isinstance(self.active_sheet.cell(row=row, column=self.date_column).value, dt.datetime):
                last_date_row = row
        if last_date_row == -1:
            return -1
        self.active_sheet.insert_rows(last_date_row + 1)
        cell = self.active_sheet.cell(row=last_date_row + 1, column=self.date_column)
        cell.value = date
        cell.number_format = "m/d/yyyy"
        return 0

    #create rows for a new month
    def new_month(self, month: int, year: int):
        start_row: int = self.title_row + 1
        for row in range(1, self.active_sheet.max_row + 1):
            value = self.active_sheet.cell(row=row, column=self.date_column).value
            if isinstance(value, dt.datetime) and \
                (value.month == month or value.month == month - 1):
                start_row = row
        end_day: int = monthrange(year, month)[1]
        if self.active_sheet.cell(row=start_row, column=self.date_column).value.month == month - 1:
            start_day: int = 1
            day: int = start_day
            row: int = start_row + 1
        else:
            start_day: int = self.active_sheet.cell(row=start_row, column=self.date_column).value.day + 1
            day: int = start_day
            row: int = start_row + 1
        while day <= end_day:
            self.active_sheet.insert_rows(row)
            cell = self.active_sheet.cell(row=row, column=self.date_column)
            cell.value = dt.datetime(year, month, day)
            cell.number_format = "m/d/yyyy"
            day += 1
            row += 1
        return 0

    #add hours
    def add_hours(self, date: dt.date, customer: str, engineer: str, hours: int):
        engineer_column: int = -1
        for column in self.engineer_columns:
            if self.active_sheet[column + str(self.title_row)].value == engineer:
                engineer_column = pyxl.utils.column_index_from_string(column)
                break
        if engineer_column == -1:
            return -1
        customer_column: int = -1
        if self.engineer_columns.index(pyxl.utils.get_column_letter(engineer_column)) == 0:
            beginning_index: int = self.date_column + 1
        else:
            beginning_index: int = pyxl.utils.column_index_from_string(self.engineer_columns[self.engineer_columns.index(engineer) - 1]) + 1
        for column in range(beginning_index, engineer_column):
            if self.active_sheet.cell(row=self.title_row, column=column).value == customer:
                customer_column = column
                break
        if customer_column == -1:
            return -1
        row: int = self.row_from_date(date)
        print(row)
        self.active_sheet.cell(row=row, column=customer_column).value = hours
        return 0

if __name__ == "__main__":
    engineer_columns: list = ["I", "Q", "Y"]
    sheet = SpreadsheetHandler("template.xlsx", engineer_columns)
    #print(sheet.engineer_names())
    #print(sheet.customer_names())
    #print(sheet.row_from_date(dt.datetime(2018, 6, 10)))
    #print(sheet.new_row(dt.datetime(2018, 6, 11)))
    #print(sheet.add_hours(dt.datetime(2018, 6, 11), "HP", "Engineer 1", 5))
    print(sheet.new_month(6, 2018))
    sheet.save()
