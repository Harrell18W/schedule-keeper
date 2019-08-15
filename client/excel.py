import datetime as dt
import openpyxl as pyxl
import os
from shutil import copy2
import time


class Spreadsheet(object):

    def __init__(self, filename):
        self.filename = filename
        self.workbook = pyxl.load_workbook(filename)
        active_year = dt.datetime(2000, 1, 1).now().strftime('%Y')
        self.active_sheet = self.workbook[active_year]

    def save(self):
        self.workbook.save(self.filename)

    def _already_backedup(self, backup_dir):
        mtime = os.path.getmtime(self.filename)
        for file in os.listdir(backup_dir):
            if mtime == os.path.getmtime(backup_dir + file):
                return True
        return False

    def _clean_backups(self, backup_dir):
        now = time.time()
        for file in os.listdir(backup_dir):
            age = now - os.path.getmtime(backup_dir + file)
            if age > 7 * 86400:
                os.remove(backup_dir + file)

    def backup(self, backup_dir, force=False, clean=True):
        if force or not self._already_backedup(backup_dir):
            filename_base = self.filename.split('/')[-1]
            filename_base, ext = filename_base.split('.')
            timestamp = dt.datetime.now().strftime('%m-%d-%y %H-%M-%S')
            dest_path = '%s%s Backup %s.%s' % (backup_dir, filename_base,
                                               timestamp, ext)
            copy2(self.filename, dest_path)

        if clean:
            self._clean_backups(backup_dir)

    def _get_customer_column(self, customer):
        for column in range(1, self.active_sheet.max_column + 1):
            value = self.active_sheet.cell(row=1, column=column).value
            if value == customer:
                return column

    def _get_date_row(self, date):
        for row in range(1, self.active_sheet.max_row + 1):
            value = self.active_sheet.cell(row=row, column=1).value
            if ((isinstance(value, dt.date) or
                    isinstance(value, dt.datetime)) and
                    value.year == date.year and
                    value.month == date.month and
                    value.day == date.day):
                return row

    def insert_job(self, date, customer, hours):
        column = self._get_customer_column(customer)
        if not column:
            return False, 'Customer %s not found in spreadsheet.' % customer
        row = self._get_date_row(date)
        if not row:
            return False, 'Date %s not found in spreadsheet.' \
                % date.strftime('%b %d %Y')
        cell = self.active_sheet.cell(row=row, column=column)
        column_letter = pyxl.utils.get_column_letter(column)
        try:
            cell.value = cell.value + hours if cell.value else hours
        except TypeError:
            return False, 'Invalid data exists in cell ' \
                '%s%d.' % (column_letter, row)
        return True, '%s, %s, %.2f hours (%s%d)' \
            % (customer, date.strftime('%b %d %Y'), hours, column_letter, row)
